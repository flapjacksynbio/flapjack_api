import os
import datetime
import openpyxl as opxl
import numpy as np
import pandas as pd
from itertools import islice
from .models import *


def synergy_get_signal_names(ws):
    """
    Params
    - ws: openpyxl worksheet object
    Returns
    - signals: List with the signal names
    """
    # first column as Cell objects
    first_col = ws['A']
    # first column as values
    first_col_arr = np.array([c.value for c in first_col])
    # index where signals information starts
    start_index = np.where(first_col_arr=='End Kinetic')[0][0]
    # obtain signals names, including the "Result" cell, if exists
    signal_names = [val for val in first_col_arr[start_index+1:] if val != None]
    return signal_names

def bmg_get_signal_names(ws_od, ws_fluo):
    """
    Params
    - ws_od: openpyxl worksheet object for od measurements
    - ws_fluo: openpyxl worksheet object for fluo measurements
    Returns
    - signals: List with the signal names
    """
    # OD
    #first column as Cell objects
    first_col_od = ws_od['1']
    # first column as values
    first_col_arr_od = np.array([c.value for c in first_col_od])
    sig_un_od = list(np.unique(first_col_arr_od[2:]))

    # Fluo
    first_col_fluo = ws_fluo['1']
    first_col_arr_fluo = np.array([c.value for c in first_col_fluo])
    sig_un_fluo = list(np.unique(first_col_arr_fluo[2:]))

    # Joined signals
    signal_names = sig_un_od + sig_un_fluo
    return signal_names


def synergy_rows_list(ws, signal_names):
    """
    Params
    - ws: openpyxl worksheet object
    - signal_names: name given to the signal by the user, i.e., "CFP"
    Returns
    - rows: list of tuples, each containing ('signal', signal_row_position)
    """
    rows = [(celda.value, celda.row)
                  for celda in ws['A']
                  if celda.value in signal_names]
    return rows


def synergy_clean_data(signal_names, df, rows):
    """
    Params
    - signal_names: name given to the signal by the user, i.e., "CFP"
    - df: DataFrame containg all data from sheet
    - rows: list of tuples, each containing ('signal', signal_row_position)
    Returs
    - dfs: DataFrame cleaned of data which are not part of the measurements
    """
    dfs = {}
    for i in range(len(rows)):
        if i == 0:
            df2 = pd.DataFrame(df.iloc[0:rows[i][1] - 3])
        else:
            df2 = pd.DataFrame(df.iloc[rows[i-1][1] + 1:rows[i][1] - 3])
        synergy_fix_time(df2)
        df2.index = range(len(df2['Time']))
        dfs[signal_names[i]] = df2
    return dfs


# Changes time's format from datetime to fraction
def synergy_fix_time(df):
    t = np.array([])
    for i, value in enumerate(df['Time']):
        if value == datetime.datetime(1899, 12, 30):
            value = datetime.time(0, 0, 0)
        try:
            t = np.append(t, value.day*24 + value.hour + value.minute/60 + value.second/3600)
        except:
            t = np.append(t, value.hour + value.minute/60 + value.second/3600)
    df['Time'] = t

"""
def bmg_fix_time(df):
    time = []
    for t in df.index:
        time_string = t
        try:
            datetime_obj = datetime.datetime.strptime(time_string, '%H h %M min')
        except:
            datetime_obj = datetime.datetime.strptime(time_string, '%H h ')
        total_mins = (datetime_obj.hour * 60) + datetime_obj.minute
        time_in_hours = total_mins / 60
        time.append(time_in_hours)
    df = df.reset_index(drop=True)
    df['Time'] = time
    return df
"""
def bmg_fix_time(df):
    time = []
    for t in df.index:   
        total_mins = int(t.split('h')[0])
        if 'm' in t:
            total_mins += int(t.split('h')[1].split('min')[0])
        time_in_hours = total_mins / 60
        time.append(time_in_hours)
    df = df.reset_index(drop=True)
    df['Time'] = time
    return df

def get_all_tables(ws, columns):
    length = len(ws['A'])
    ntables = (length + 1) // 10
    table_list = []
    names_list = []
    for i in range(ntables):
        name,values = table_values(ws, i*10+1, columns)
        table_list.append(values)
        names_list.append(name)
    return names_list, table_list


def table_values(ws, row, columns):
    name = ws[row][0].value
    vals = []
    for cell_row in range(row + 1, row + 9):
        for cell_col in range (1, 13):
            val = ws[cell_row][cell_col].value
            if not val:
                vals.append(0)
            else:
                vals.append(val)
    return name, dict(zip(columns, vals))


def synergy_load_data(ws, signal_names, signal_map):
    rows_ini = synergy_rows_list(ws, signal_names)
    ws.delete_rows(0, rows_ini[0][1] + 1)
    rows = synergy_rows_list(ws, signal_names)
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, columns=cols)

    names = [signal_map[rows_ini[i][0]] for i in range(len(rows_ini)-1)]
    dfs = synergy_clean_data(names, df, rows)
    return dfs

def bmg_load_data(wb, signal_map, columns):
    sheet_od = pd.DataFrame(wb['OD'].values)
    sheet_od.columns = sheet_od.iloc[0]
    sheet_od = sheet_od[1:]

    sheet_od = sheet_od.drop(['Content'],axis=1)
    columns_od = sheet_od.Well.dropna()
    sheet_od = sheet_od.drop(['Well'],axis=1)
    columns_od = list(columns_od.values)
    columns_od = [x[0]+str(int(x[1])) for x in columns_od]

    sheet_od = sheet_od.T
    sheet_od = sheet_od.set_index(1)
    sheet_od.columns = columns_od
    sheet_od = bmg_fix_time(sheet_od)

    sheet_fluo = pd.DataFrame(wb['Fluo'].values)
    sheet_fluo.columns = sheet_fluo.iloc[0]
    sheet_fluo = sheet_fluo[1:]

    sheet_fluo = sheet_fluo.drop(['Content'],axis=1)
    columns_fluo = sheet_fluo.Well.dropna()
    sheet_fluo = sheet_fluo.drop(['Well'],axis=1)
    columns_fluo = list(columns_fluo.values)
    columns_fluo = [x[0]+str(int(x[1])) for x in columns_fluo]

    sheet_fluo = sheet_fluo.T
    sigs = []
    for i in sheet_fluo.index:
        for s in signal_map:
            if s in i:
                sigs.append(signal_map[s])

    sheet_fluo = sheet_fluo.reset_index(drop=True)
    sheet_fluo = sheet_fluo.dropna(axis=1)
    sheet_fluo = sheet_fluo.set_index(1)
    sheet_fluo.columns = columns_fluo
    sheet_fluo = bmg_fix_time(sheet_fluo)
    sheet_fluo['Fluo'] = sigs

    dfs = {}
    dfs['OD'] = sheet_od
    for s in np.unique(sheet_fluo['Fluo']):
        dfs[s] = sheet_fluo[sheet_fluo.Fluo==s].drop('Fluo', axis=1)
        
    return dfs

# Also works for BMG
def synergy_load_meta(wb, columns):
    meta_dict = {}
    ws_names = ['Strains', 'Media', 'DNA', 'Chemicals']
    for ws_name in ws_names:
        if ws_name in ['Strains', 'Media']:
            name, dict_ws = table_values(wb[ws_name], 1, columns)
            meta_dict[name] = dict_ws
        elif ws_name == 'DNA':
            names, dicts_ws = get_all_tables(wb[ws_name], columns)
            for i in range(len(names)):
                meta_dict[names[i]] = dicts_ws[i]
        elif ws_name == 'Chemicals':
            if 'Chemicals' in wb.sheetnames:
                names, dicts_ws = get_all_tables(wb[ws_name], columns)
                for i in range(len(names)):
                    meta_dict[names[i]+' chemical'] = dicts_ws[i]
    return pd.DataFrame(meta_dict).transpose()
